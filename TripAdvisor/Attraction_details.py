from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Output Excel file path
path = r'D:\attraction_details.xlsx'

# Load the workbook containing attraction links
excel_path = r'D:\attraction_links.xlsx'
wb_links = openpyxl.load_workbook(excel_path)
sh_links = wb_links.active

# Get the links from the second column
attraction_urls = [row[1] for row in sh_links.iter_rows(min_row=2, max_col=2, values_only=True)]

# Initialize Chrome WebDriver with options to ignore loading images and block JavaScript
chrome_options = Options()
chrome_options.add_experimental_option(
    "prefs", {
        # block image loading
        "profile.managed_default_content_settings.images": 2,
        # block JavaScript
        'profile.managed_default_content_settings.javascript': 2,
    }
)
chrome_options.add_experimental_option("detach", True)
chrome_options.binary_location = './/chrome-win64/chrome.exe'
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.maximize_window()

# Create a new workbook for storing attraction details
wb_details = openpyxl.Workbook()
sh_details = wb_details.active
sh_details.append(['ATTRACTION_NAME', 'CATEGORY', 'CONTINENT', 'COUNTRY', 'PROVINCE', 'CITY', 'RATING', 'TIMING'])  # Adding headers to the Excel file

# Iterate through each attraction URL
for url in attraction_urls:
    driver.get(url)

    attraction_name = WebDriverWait(driver, 260).until(
        EC.presence_of_all_elements_located((By.XPATH, '//div[@class="iSVKr"]'))
    )

    try:
        attraction_category = WebDriverWait(driver, 1).until(
        EC.presence_of_all_elements_located((By.XPATH, '//div[@class="zCoYj"]'))
    )
    except:
        attraction_category = ['N/A'] * len(attraction_name)

    attraction_continent = WebDriverWait(driver, 1).until(
        EC.presence_of_all_elements_located((By.XPATH, '(//div[@class="KCGqk _T Cj"])[1]'))
    )
    attraction_country = WebDriverWait(driver, 1).until(
        EC.presence_of_all_elements_located((By.XPATH, '((//div[@class="KCGqk _T Cj"])[2])'))
    )
    attraction_province = WebDriverWait(driver, 1).until(
        EC.presence_of_all_elements_located((By.XPATH, '((//div[@class="KCGqk _T Cj"])[3])'))
    )
    attraction_city = WebDriverWait(driver, 1).until(
        EC.presence_of_all_elements_located((By.XPATH, '((//div[@class="KCGqk _T Cj"])[4])'))
    )

    try:
        attraction_rating = WebDriverWait(driver, 1).until(
            EC.presence_of_all_elements_located((By.XPATH, '//div[@class="biGQs _P fiohW hzzSG uuBRH"]'))
        )
    except:
        attraction_rating = ['N/A'] * len(attraction_name)

    try:
        attraction_timing = WebDriverWait(driver, 1).until(
            EC.presence_of_all_elements_located((By.XPATH, '//span[@class="EFKKt"]'))
        )
    except:
        attraction_timing = ['N/A'] * len(attraction_name)

    # Extract and save information for each attraction
    for name, category, continent, country, province, city, rating, timing in zip(attraction_name, attraction_category, attraction_continent, attraction_country
                                                            ,attraction_province, attraction_city, attraction_rating, attraction_timing):
        names = name.text if name != 'N/A' else 'N/A'
        categories = category.text if category != 'N/A' else 'N/A'
        continents = continent.text if continent else 'N/A'
        countries = country.text if country else 'N/A'
        provinces = province.text if province else 'N/A'
        cities = city.text if city else 'N/A'
        ratings = rating.text if rating != 'N/A' else 'N/A'
        timings = timing.text if timing != 'N/A' else 'N/A'
        sh_details.append([names, categories, continents, countries, provinces, cities, ratings, timings])

# Save the workbook for attraction details
wb_details.save(path)

# Close the workbooks and WebDriver when done
wb_links.close()
wb_details.close()
driver.quit()
